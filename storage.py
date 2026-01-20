import os
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelStore:
    """
    Represents how inputs from the user will be stored.
    Attributes:
        filename (str): The name of the excel file where data is stored.
    """
    def __init__(self, filename: str):
        """
        Initializes the filename.

        Arguments:
            filename (str): The name of the excel file where data is stored.
        """
        self.filename = filename

    def _sheet_exists(self, sheet_name: str) -> bool:
        """
        Validates that the sheet exists.
        
        Arguments:
            sheet_name (str): The specific sheet that the user is trying to access. 
        """
        if not os.path.exists(self.filename):
            return False
        try:
            pd.read_excel(self.filename, sheet_name=sheet_name, nrows=1)
            return True
        except Exception:
            return False

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Uses Pandas read_excel to return a Pandas DataFrame.

        Arguments:
            sheet_name (str): The specific sheet that the user is trying to access.

        Returns:
            pd.DataFrame: The sheet of an excel file that is turned into a Pandas DataFrame.
        """
        if not os.path.exists(self.filename):
            return pd.DataFrame()
        try:
            return pd.read_excel(self.filename, sheet_name=sheet_name)
        except Exception:
            return pd.DataFrame()

    def append_rows(self, sheet_name: str, columns: List[str], rows: List[Dict[str, Any]]) -> None:
        """
        Uses Pandas and Openpyxl to add rows of data to the specific sheet in the excel file.

        Arguments:
            sheet_name (str): The specific sheet the user is trying to add data to.
            columns (List[str]): A list of headers for the sheet.
            rows (List[Dict[str, Any]]): A list of a dictionary including all of the current rows.
        """
        new_df = pd.DataFrame(rows, columns=columns)

        file_exists = os.path.exists(self.filename)
        sheet_exists = self._sheet_exists(sheet_name)

        if not file_exists:
            # First time: let pandas create file + sheet
            with pd.ExcelWriter(self.filename, engine="openpyxl", mode="w") as writer:
                new_df.to_excel(writer, sheet_name=sheet_name, index=False)
            return

        # File exists → use openpyxl to preserve formatting
        wb = load_workbook(self.filename)

        if sheet_exists:
            ws = wb[sheet_name]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(sheet_name)
            start_row = 1

        for r_idx, row in enumerate(
            dataframe_to_rows(new_df, index=False, header=not sheet_exists),
            start=start_row
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(self.filename)
        
    def get_last_entries(self, sheet_name: str, n: int = 5) -> None:
        """
        Gets the last 5 entries of a sheet.

        Arguments:
            sheet_name (str): The specific sheet the user is trying to add data to.
            n (int) = 5: Last 5 entries in the sheet.
        """
        df = self.read_sheet(sheet_name)

        if df.empty:
            print(f"❗ No entries found in '{sheet_name}'.")
            return

        if "Date" not in df.columns:
            print(f"❌ Sheet '{sheet_name}' missing required 'Date' column.")
            return

        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

        recent = (
            df
            .dropna(subset=["Date"])
            .sort_values("Date")
            .tail(n)
        )

        if recent.empty:
            print("❗ No valid dated entries found.")
            return

        print(f"\n📋 Last {n} entries from '{sheet_name}':\n")
        print(recent.to_string(index=False))

        

        

    